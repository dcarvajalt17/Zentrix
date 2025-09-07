import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

# ===== LISTAS DE GESTIÓN FORZADA =====
FORZAR_BUFFER = {
    '901007','901012','901013','410014','410018','410019','410020','410021','410024','410056',
    '410060','410061','410068','210003','210012','210016','210019','210020','210021','310001',
    '310004','310006','520011','520012','520019','520020','520025','520026','520031','520032',
    '520033','520034','520029','520030','520040','520050','520178','520060','520123','520164',
    '820002','710018','710019','710023','710028'
}

FORZAR_NO_BUFFER = {
    '701006','701008','710002','710003','710004','710008','710009','710013','710015','710035',
    '710036','410032','410039','410047','410053','410055','410069','410070','410001','410002',
    '410011','410051','410052','410057','410064','410065','410066','410067','410071','410050',
    '410075','520186','210002','210005','210006','210007','210011','210013','210017','210022',
    '310002','310003','310007','520051','520103','520127','520128','520130','520043','520045',
    '520166','520167','520039','520118','520129','520132','520169','520170','520171','520172',
    '520173','520174','520176','520180','520181','520182','520134','520138','520140','520146',
    '520161','520168','520175','520177','520155','520156','520157','520063','520064','520122',
    '520165','440057','440059','440061','440062','440063','440064','440065','410054','520136',
    '990004','990014','990027','990028','990034','990012','410034','520115','210015','520117',
    '820001','820003','820004','820074','820078','820079','710025','710027'
}

# ===== PARÁMETROS DE CONTROL =====
APLICAR_ADU_MINIMO = True
REFORZAR_PICOS = True
FRACCIONAR_PEDIDOS = True
COBERTURA_NO_BUFFER = 1.25
LIMITE_FACTOR_PROYECCION = (0.6, 1.8)   # puedes ampliar a (0.3, 3.0) si tu estacionalidad es fuerte

# ===================== Utilidades =====================

def normalizar_codigo(x):
    if pd.isna(x):
        return ''
    if isinstance(x, (int, float)):
        return str(int(x)).strip().upper()
    return str(x).strip().upper()

def calcular_ltf(lead_time_dias):
    if pd.isnull(lead_time_dias):
        return 1.0
    if lead_time_dias <= 5:
        return 0.3
    elif lead_time_dias <= 15:
        return 0.5
    elif lead_time_dias <= 30:
        return 0.75
    else:
        return 1.0

# ====== Fórmula estándar DDMRP ======
def calcular_buffer_ddmrp(row):
    """
    DDMRP:
      Yellow = ADU * DLT
      Red    = LTF * ADU * DLT  +  Variabilidad (VF * Red_base)
      Green  = max(MOQ, order_cycle * ADU)   # SIN LTF ni DLT
      TO_Yellow = Red_Total + Yellow
      TO_Green  = TO_Yellow + Green
    """
    adu = float(row['ADU']) * float(row.get('projection_factor', 1))
    adu = max(0.0, adu)

    dlt = max(1.0, float(row['DLT']))
    ltf = max(0.0, float(row['LTF']))
    vf  = max(0.0, float(row.get('VF', 0.0)))
    moq = max(0.0, float(row.get('MOQ', 0.0)))
    oc  = max(1.0, float(row.get('order_cycle', 1)))

    yellow = adu * dlt
    red_base   = ltf * adu * dlt
    red_safety = red_base * vf
    red_total  = red_base + red_safety
    green = max(moq, oc * adu)

    to_yellow = red_total + yellow
    to_green  = to_yellow + green

    return pd.Series({
        'Yellow': yellow,
        'Green': green,
        'Red_Total': red_total,
        'TO_Yellow': to_yellow,
        'TO_Green': to_green,
        'Avg_IP': red_total + green / 2.0
    })

# ====== Estacionalidad semanal desde histórico ======

def _semana_del_mes_from_day(day: int) -> int:
    """Devuelve 1..4 a partir del día del mes (aprox. semanas de 7 días)."""
    s = (int(day) - 1) // 7 + 1
    return int(min(max(s, 1), 4))

def construir_perfiles_estacionales(df_hist, df_dim):
    """
    Construye perfiles {ref: [w1,w2,w3,w4]} a partir del histórico real.
    Si no hay histórico por referencia, cae a Familia/Subfamilia; luego a uniforme.
    df_hist: columnas -> ['Referencia' o 'Codigo','Fecha','Consumo']
    df_dim : índice 'Codigo' con ['Familia','Subfamilia']
    """
    df_hist = df_hist.copy()
    # renombrar si viene como 'Codigo'
    if 'Referencia' not in df_hist.columns and 'Codigo' in df_hist.columns:
        df_hist = df_hist.rename(columns={'Codigo': 'Referencia'})

    if df_hist.empty:
        # si no hay histórico, devolvemos siempre uniforme
        def get_perfil(_, __, ___): return [0.25, 0.25, 0.25, 0.25]
        return get_perfil

    df_hist['Referencia'] = df_hist['Referencia'].apply(normalizar_codigo)
    df_hist['Fecha'] = pd.to_datetime(df_hist['Fecha'], errors='coerce')
    df_hist['Consumo'] = pd.to_numeric(df_hist['Consumo'], errors='coerce').fillna(0)
    df_hist = df_hist.dropna(subset=['Referencia','Fecha'])

    df_hist['SemanaMes'] = df_hist['Fecha'].dt.day.apply(_semana_del_mes_from_day)

    # Map a familia/subfamilia
    df_hist = df_hist.merge(
        df_dim.reset_index().rename(columns={'Codigo': 'Referencia'}),
        on='Referencia', how='left'
    )

    # Perfil por referencia
    ref_sum = df_hist.groupby(['Referencia','SemanaMes'], as_index=False)['Consumo'].sum()
    ref_total = ref_sum.groupby('Referencia', as_index=False)['Consumo'].sum().rename(columns={'Consumo':'Total'})
    ref_pf = ref_sum.merge(ref_total, on='Referencia', how='left')
    ref_pf['Peso'] = ref_pf['Consumo'] / ref_pf['Total'].replace(0, pd.NA)
    ref_pf = ref_pf.dropna(subset=['Peso'])

    perfiles_ref = {
        ref: [float(ref_pf[(ref_pf['Referencia']==ref) & (ref_pf['SemanaMes']==w)]['Peso'].sum()) for w in [1,2,3,4]]
        for ref in ref_pf['Referencia'].unique()
    }

    # Perfil por familia/subfamilia (fallback)
    fam_sum = df_hist.groupby(['Familia','Subfamilia','SemanaMes'], as_index=False)['Consumo'].sum()
    fam_total = fam_sum.groupby(['Familia','Subfamilia'], as_index=False)['Consumo'].sum().rename(columns={'Consumo':'Total'})
    fam_pf = fam_sum.merge(fam_total, on=['Familia','Subfamilia'], how='left')
    fam_pf['Peso'] = fam_pf['Consumo'] / fam_pf['Total'].replace(0, pd.NA)
    fam_pf = fam_pf.dropna(subset=['Peso'])

    perfiles_fam = {
        (f,s): [float(fam_pf[(fam_pf['Familia']==f) & (fam_pf['Subfamilia']==s) & (fam_pf['SemanaMes']==w)]['Peso'].sum())
                for w in [1,2,3,4]]
        for f,s in fam_pf[['Familia','Subfamilia']].drop_duplicates().itertuples(index=False, name=None)
    }

    def get_perfil(ref, fam, subfam):
        base = perfiles_ref.get(ref)
        if base and sum(base) > 0:
            tot = sum(base)
            return [x/tot for x in base]
        base = perfiles_fam.get((fam, subfam))
        if base and sum(base) > 0:
            tot = sum(base)
            return [x/tot for x in base]
        return [0.25, 0.25, 0.25, 0.25]

    return get_perfil

# ===================== Núcleo =====================

def ejecutar_completo(ruta_excel, ruta_maestro):
    # Parámetros
    df_params = pd.read_excel(ruta_excel, sheet_name='Referencias')
    df_params['LTF'] = df_params['DLT'].apply(calcular_ltf)

    # Proyección (si existe)
    xls_in = pd.ExcelFile(ruta_excel)
    df_proj = pd.read_excel(ruta_excel, sheet_name='Proyeccion') if 'Proyeccion' in xls_in.sheet_names else None

    # OC Pendientes
    df_oc = pd.read_excel(ruta_excel, sheet_name='OC Pendientes')

    # Dimensiones (familia/subfamilia/desc)
    df_dim_all = pd.read_excel(ruta_maestro, sheet_name='Codigo')
    df_dim_all['Codigo'] = df_dim_all['Codigo'].apply(normalizar_codigo)
    df_dim = df_dim_all.set_index('Codigo')[['Familia','Subfamilia']]
    df_desc = df_dim_all.set_index('Codigo')[['Familia','Subfamilia','Descripcion_F']]

    # Histórico para perfiles
    try:
        df_hist = pd.read_excel(ruta_maestro, sheet_name='Consumo')  # columnas: Codigo/Referencia, Fecha, Consumo
    except Exception:
        df_hist = pd.DataFrame(columns=['Referencia','Fecha','Consumo'])
    get_perfil = construir_perfiles_estacionales(df_hist, df_dim)

    # Normalizar
    df_params['Referencia'] = df_params['Referencia'].apply(normalizar_codigo)
    if df_proj is not None:
        df_proj['Referencia'] = df_proj['Referencia'].apply(normalizar_codigo)
    df_oc['Material'] = df_oc['Material'].apply(normalizar_codigo)
    df_oc['Fecha de entrega'] = pd.to_datetime(df_oc['Fecha de entrega'], errors='coerce')

    # Inventarios (openpyxl para respetar data_only/formulas)
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb['Inventario']
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    headers = data[0]
    df_inv = pd.DataFrame(data[1:], columns=headers)
    df_inv['Material'] = df_inv['Material'].apply(normalizar_codigo)
    df_inv['precio'] = pd.to_numeric(df_inv['precio'], errors='coerce')
    df_inv['Libre utilización'] = pd.to_numeric(df_inv['Libre utilización'], errors='coerce')

    df_inv_agg = df_inv.groupby('Material', as_index=False).agg({
        'Libre utilización': 'sum',
        'precio': 'mean'
    })
    inv_map = df_inv_agg.set_index('Material')['Libre utilización'].to_dict()
    precio_map = df_inv_agg.set_index('Material')['precio'].to_dict()

    # Fechas desde proyección (ej. columnas '2025-09', '2025-10', ...)
    fechas = []
    if df_proj is not None:
        fechas = [(pd.to_datetime(col, errors='coerce'), col) for col in df_proj.columns if '20' in str(col)]
        fechas = [(d, c) for d, c in fechas if pd.notna(d)]
        fechas.sort()

    resultados = {}

    for ref in df_params['Referencia']:
        historico = []
        params = df_params[df_params['Referencia'] == ref].iloc[0].copy()

        # Determinar gestión priorizando listas forzadas
        if ref in FORZAR_BUFFER:
            gestion = 'Buffer'
        elif ref in FORZAR_NO_BUFFER:
            gestion = 'No Buffer'
        else:
            gestion = params['gestion']

        if gestion not in ['Buffer', 'No Buffer']:
            continue

        adu = params['ADU']
        stock_inicial = inv_map.get(ref, 0.0)
        precio = precio_map.get(ref, 0.0)
        pedidos_futuros = []

        # Perfil estacional (ref → familia/subfamilia → uniforme)
        fam = df_dim.at[ref, 'Familia'] if ref in df_dim.index else None
        subfam = df_dim.at[ref, 'Subfamilia'] if ref in df_dim.index else None
        perfil_sem = get_perfil(ref, fam, subfam)  # 4 pesos que suman 1

        # Si no hay proyección, usar un único mes ficticio
        fechas_iter = fechas if fechas else [(datetime.today(), None)]

        for fecha, col in fechas_iter:
            consumo_mes = 0.0
            if df_proj is not None and col is not None and ref in df_proj['Referencia'].values:
                consumo_mes = df_proj.set_index('Referencia').at[ref, col]
            consumo_mes = float(0.0 if pd.isna(consumo_mes) else consumo_mes)
            dias_semana = 7

            for semana in range(1, 5):
                # Consumo semanal con perfil estacional
                consumo_semana = consumo_mes * perfil_sem[semana - 1]

                fecha_eval = fecha + timedelta(days=(semana - 1) * dias_semana)
                fecha_fin  = fecha_eval + timedelta(days=dias_semana)

                recepciones_oc = df_oc[(df_oc['Material'] == ref) &
                                       (df_oc['Fecha de entrega'] > fecha_eval) &
                                       (df_oc['Fecha de entrega'] <= fecha_fin)]['Por entregar (cantidad)'].sum()
                recep_pedido = sum(p for (fe, p) in pedidos_futuros if fecha_eval < fe <= fecha_fin)
                recepciones = recepciones_oc + recep_pedido

                transito_oc = df_oc[(df_oc['Material'] == ref) & (df_oc['Fecha de entrega'] > fecha_fin)]['Por entregar (cantidad)'].sum()
                transito_ped = sum(p for (fe, p) in pedidos_futuros if fe > fecha_fin)
                transito = transito_oc + transito_ped

                # Posición/NFP = stock + recepciones (<= fin) + tránsito (> fin)
                posicion = stock_inicial + recepciones + transito

                # ADU mínimo si corresponde
                if APLICAR_ADU_MINIMO and (adu == 0 or adu < params['MOQ'] / max(params['DLT'], 1)):
                    adu = max(adu, params['MOQ'] / max(params['DLT'], 1))

                consumo_proy_futuro = consumo_mes  # total del mes (para refuerzos)

                # Factor de proyección para bandas por mes
                base_mes = max(adu * 4.0, 1e-9)
                consumo_mes_pf = base_mes if (df_proj is None or col is None) else float(consumo_mes)
                pf = consumo_mes_pf / base_mes
                pf = max(LIMITE_FACTOR_PROYECCION[0], min(LIMITE_FACTOR_PROYECCION[1], pf))
                params_loc = params.copy()
                params_loc['projection_factor'] = pf

                pedido = 0.0
                fecha_entrega = fecha_eval + timedelta(days=int(params['DLT']))

                if gestion == 'Buffer':
                    buf = calcular_buffer_ddmrp(params_loc)
                    # Disparo clásico: NFP < TO_Yellow ⇒ pedir hasta TO_Green
                    if posicion < buf['TO_Yellow']:
                        pedido = max(buf['TO_Green'] - posicion, 0)
                        fecha_entrega = fecha_eval + timedelta(days=int(params['DLT']))
                        if FRACCIONAR_PEDIDOS and pedido > 2 * buf['Green']:
                            mitad = pedido / 2.0
                            pedidos_futuros.append((fecha_entrega, mitad))
                            pedidos_futuros.append((fecha_entrega + timedelta(days=int(params['DLT'])), mitad))
                            pedido = 0.0
                else:  # No Buffer
                    inv_min = max(
                        adu * params['DLT'] * (1 + params['VF']),
                        adu * params['DLT'] * COBERTURA_NO_BUFFER
                    )
                    inv_max = inv_min + adu * params['DLT']
                    inv_objetivo = (inv_min + inv_max) / 2
                    if REFORZAR_PICOS and consumo_proy_futuro > 2 * adu:
                        inv_min = max(inv_min, consumo_proy_futuro)
                    if posicion < inv_objetivo:
                        pedido = max(inv_max - posicion, 0)
                        fecha_entrega = fecha_eval + timedelta(days=int(params['DLT']))
                        if FRACCIONAR_PEDIDOS and pedido > 2 * inv_max:
                            mitad = pedido / 2.0
                            pedidos_futuros.append((fecha_entrega, mitad))
                            pedidos_futuros.append((fecha_entrega + timedelta(days=int(params['DLT'])), mitad))
                            pedido = 0.0

                if pedido > 0:
                    pedidos_futuros.append((fecha_entrega, pedido))

                # Inventario proyectado de cierre de semana:
                inventario_proy = max(0.0, stock_inicial + recepciones - consumo_semana)
                cobertura = inventario_proy / adu if adu else 0

                entry = {
                    'Referencia': ref,
                    'Mes': fecha.strftime('%b-%y'),
                    'Semana': semana,
                    'Fecha_Evaluacion': fecha_eval.strftime('%Y-%m-%d'),
                    'Recepciones_Planeadas': round(recepciones, 2),
                    'Transito': round(transito, 2),
                    'Stock_Inicial': round(stock_inicial, 2),
                    'Posicion': round(posicion, 2),
                    'Fecha_Pedido': fecha_eval.strftime('%Y-%m-%d'),
                    'Fecha_Entrega_Pedido': fecha_entrega.strftime('%Y-%m-%d'),
                    'Consumo_Proy': round(consumo_semana, 2),
                    'Cantidad_Pedir': round(pedido, 2),
                    'Inventario_Proy': round(inventario_proy, 2),
                    'Nivel_Cobertura_Dias': round(cobertura, 2),
                    'precio': round(precio, 2),
                    'vr_compra': round(precio * pedido, 2)
                }

                if gestion == 'Buffer':
                    entry.update(calcular_buffer_ddmrp(params_loc).to_dict())
                else:
                    entry.update({
                        'Inventario_Minimo': round(inv_min, 2),
                        'Inventario_Maximo': round(inv_max, 2),
                        'Inventario_Objetivo': round(inv_objetivo, 2)
                    })

                historico.append(entry)
                stock_inicial = inventario_proy

        resultados[ref] = (gestion, pd.DataFrame(historico))

    return df_params.set_index('Referencia'), resultados

# ===================== Exportador =====================

def exportar_resumen(ruta_excel, ruta_maestro, salida):
    df_params, resultados = ejecutar_completo(ruta_excel, ruta_maestro)

    # Dimensiones y descripción para enriquecer hojas
    df_maestro = pd.read_excel(ruta_maestro, sheet_name='Codigo')
    df_maestro['Codigo'] = df_maestro['Codigo'].apply(normalizar_codigo)
    df_maestro = df_maestro.set_index('Codigo')[['Familia','Subfamilia','Descripcion_F']]

    wb = Workbook()
    wb.remove(wb.active)

    for categoria in ['Buffer', 'No Buffer']:
        datos = [df for _, (gest, df) in resultados.items() if gest == categoria]
        if not datos:
            continue
        df_cat = pd.concat(datos, ignore_index=True)

        # Añadir dimensiones desde el maestro
        dim_map = df_maestro.to_dict(orient='index')
        df_cat[['Familia','Subfamilia','Descripcion_F']] = (
            df_cat['Referencia'].map(dim_map).apply(pd.Series)
        )

        # ---- Detalle
        ws_detalle = wb.create_sheet(categoria)
        for r in dataframe_to_rows(df_cat, index=False, header=True):
            ws_detalle.append(r)

        # ---- Resumen por mes (solo compras > 0)
        df_resumen = df_cat[df_cat['vr_compra'] > 0].copy()
        df_resumen['Mes_Orden'] = pd.to_datetime(df_resumen['Mes'], format='%b-%y', errors='coerce')
        df_resumen = df_resumen.dropna(subset=['Mes_Orden'])

        ws_resumen = wb.create_sheet(f"Resumen_{categoria}")

        if df_resumen.empty:
            ws_resumen.append(['Familia','Subfamilia','Referencia','Descripcion_F','Total'])
        else:
            tabla = df_resumen.pivot_table(
                index=['Familia','Subfamilia','Referencia','Descripcion_F'],
                columns='Mes_Orden',
                values='vr_compra',
                aggfunc='sum',
                fill_value=0
            )
            tabla = tabla.sort_index(axis=1)
            if len(tabla.columns) > 0:
                tabla.columns = [d.strftime('%b-%y') for d in tabla.columns]

            tabla['Total'] = tabla.sum(axis=1)
            tabla = tabla.reset_index()

            total = pd.DataFrame(
                [['TOTAL','','',''] + list(tabla.iloc[:,4:].sum())],
                columns=tabla.columns
            )
            tabla = pd.concat([tabla, total], ignore_index=True)

            for r in dataframe_to_rows(tabla, index=False, header=True):
                ws_resumen.append(r)

    wb.save(salida)

# ===================== Runner =====================

if __name__ == '__main__':
    exportar_resumen('Referencia V2.xlsx', 'data-consumo1.xlsx', 'Resumen_Buffer_NoBuffer_Semanal.xlsx')
    print("✅ Archivo generado con DDMRP corregido + estacionalidad semanal desde histórico.")
