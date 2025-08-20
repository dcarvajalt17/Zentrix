import pandas as pd
import calendar
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

FORZAR_BUFFER = {
    '901007', '901012', '901013', '410014', '410018', '410019', '410020', '410021', '410024', '410056',
    '410060', '410061', '410068', '210003', '210012', '210016', '210019', '210020', '210021', '310001',
    '310004', '310006', '520011', '520012', '520019', '520020', '520025', '520026', '520031', '520032',
    '520033', '520034', '520029', '520030', '520040', '520050', '520178', '520060', '520123', '520164',
    '820002', '710018', '710019', '710023', '710028'
}

def normalizar_codigo(x):
    if pd.isna(x):
        return ''
    if isinstance(x, (int, float)):
        return str(int(x)).strip().upper()
    return str(x).strip().upper()

def limpiar_precio_en_excel(archivo_excel):
    wb = load_workbook(archivo_excel, data_only=True)
    ws = wb["Inventario"]
    headers = [cell.value for cell in ws[1]]
    if "precio" not in [str(h).strip().lower() for h in headers]:
        print("❌ No se encontró columna 'precio'")
        return
    col_index = [i for i, h in enumerate(headers) if str(h).strip().lower() == 'precio'][0] + 1
    for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
        for cell in row:
            if cell.data_type == 'f':
                cell.value = cell.internal_value
    wb.save(archivo_excel)
    print("✅ Columna 'precio' limpiada (fórmulas convertidas a valores)")

def calcular_ltf(lead_time_dias):
    if pd.isnull(lead_time_dias):
        return 1.0
    if lead_time_dias <= 5:
        return 0.3
    elif lead_time_dias <= 15:
        return 0.5
    elif lead_time_dias <= 30:
        return 0.75
    else:
        return 1.0

def calcular_buffer_ddmrp(row):
    adu = row['ADU'] * row.get('projection_factor', 1)
    yellow = adu * row['DLT']
    green = max(row['MOQ'], row['order_cycle'] * adu, row['LTF'] * adu * row['DLT'])
    red_base = row['LTF'] * adu * row['DLT']
    red_safety = red_base * row['VF']
    red_total = red_base + red_safety
    return pd.Series({
        'Yellow': yellow,
        'Green': green,
        'Red_Total': red_total,
        'TO_Yellow': yellow + red_total,
        'TO_Green': yellow + red_total + green,
        'Avg_IP': red_total + green / 2
    })

def ejecutar_completo(ruta_excel, ruta_maestro):
    df_params = pd.read_excel(ruta_excel, sheet_name='Referencias')
    df_params['LTF'] = df_params['DLT'].apply(calcular_ltf)
    df_proj = pd.read_excel(ruta_excel, sheet_name='Proyeccion')
    df_oc = pd.read_excel(ruta_excel, sheet_name='OC Pendientes')

    df_params['Referencia'] = df_params['Referencia'].apply(normalizar_codigo)
    df_proj['Referencia'] = df_proj['Referencia'].apply(normalizar_codigo)
    df_oc['Material'] = df_oc['Material'].apply(normalizar_codigo)
    df_oc['Fecha de entrega'] = pd.to_datetime(df_oc['Fecha de entrega'], errors='coerce')

    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb['Inventario']
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    headers = data[0]
    df_inv = pd.DataFrame(data[1:], columns=headers)
    df_inv['Material'] = df_inv['Material'].apply(normalizar_codigo)
    df_inv['precio'] = pd.to_numeric(df_inv['precio'], errors='coerce')
    df_inv['Libre utilización'] = pd.to_numeric(df_inv['Libre utilización'], errors='coerce')

    df_inv_agg = df_inv.groupby('Material', as_index=False).agg({
        'Libre utilización': 'sum',
        'precio': 'mean'
    })

    inv_map = df_inv_agg.set_index('Material')['Libre utilización'].to_dict()
    precio_map = df_inv_agg.set_index('Material')['precio'].to_dict()

    fechas = [(pd.to_datetime(col), col) for col in df_proj.columns if '20' in str(col)]
    fechas.sort()

    resultados = {}

    for ref in set(df_params['Referencia']) & set(df_proj['Referencia']):
        historico = []
        params = df_params[df_params['Referencia'] == ref].iloc[0].copy()
        gestion = 'Buffer' if ref in FORZAR_BUFFER else params['gestion']
        if gestion not in ['Buffer', 'No Buffer']:
            continue

        adu = params['ADU']
        stock_inicial = inv_map.get(ref, 0.0)
        precio = precio_map.get(ref, 0.0)
        pedidos_futuros = []
        print(f"🔍 {ref} | Desv: {params.get('Desviacion_Estandar_LT', 'No encontrada')}")

        for fecha, col in fechas:
            dias_mes = calendar.monthrange(fecha.year, fecha.month)[1]
            consumo_mes = df_proj.set_index('Referencia').at[ref, col]
            consumo_semana = (0.0 if pd.isna(consumo_mes) else consumo_mes) / 4
            dias_semana = dias_mes / 4

            for semana in range(1, 5):
                fecha_eval = datetime(fecha.year, fecha.month, 1) + timedelta(days=(semana - 1) * dias_semana)
                fecha_fin = fecha_eval + timedelta(days=dias_semana)

                recepciones_oc = df_oc[(df_oc['Material'] == ref) & 
                                       (df_oc['Fecha de entrega'] > fecha_eval) & 
                                       (df_oc['Fecha de entrega'] <= fecha_fin)]['Por entregar (cantidad)'].sum()
                recep_pedido = sum(p for (fe, p) in pedidos_futuros if fecha_eval < fe <= fecha_fin)
                recepciones = recepciones_oc + recep_pedido

                transito_oc = df_oc[(df_oc['Material'] == ref) & (df_oc['Fecha de entrega'] > fecha_fin)]['Por entregar (cantidad)'].sum()
                transito_ped = sum(p for (fe, p) in pedidos_futuros if fe > fecha_fin)
                transito = transito_oc + transito_ped

                posicion = stock_inicial + recepciones + transito
                factor = (consumo_semana / dias_semana) / adu if adu else 1
                row = params.copy()
                row['projection_factor'] = factor
                row['ADU'] = adu

                if gestion == 'Buffer':
                    buf = calcular_buffer_ddmrp(row)
                    row_base = row.copy()
                    row_base['projection_factor'] = 1
                    buf_base = calcular_buffer_ddmrp(row_base)

                    pedido = 0
                    for h in range(2, 5):
                        fecha_futura = fecha_eval + timedelta(days=dias_semana * h)
                        fecha_recepcion = fecha_futura + timedelta(days=int(row['DLT']))

                        col_adu_futuro = None
                        for fecha_c, col_fut in fechas:
                            if fecha_c >= fecha_recepcion:
                                col_adu_futuro = col_fut
                                break
                        if not col_adu_futuro:
                            col_adu_futuro = fechas[-1][1]

                        consumo_mes_futuro = df_proj.set_index('Referencia').at[ref, col_adu_futuro]
                        adu_futuro = (consumo_mes_futuro / 30) if not pd.isna(consumo_mes_futuro) else adu

                        stock_simulado = posicion - (adu_futuro * dias_semana * h)

                        if stock_simulado < buf_base['Avg_IP']:
                            pedido = max(buf_base['TO_Green'] - stock_simulado, 0)
                            fecha_entrega = fecha_eval + timedelta(days=int(row['DLT']))
                            break
                    else:
                        inv_min = adu * (1 + row['LTF'])
                        inv_max = inv_min + adu * row['DLT'] * (1 + row['VF'])
                        pedido = max(inv_max - posicion, 0)
                        fecha_entrega = fecha_eval + timedelta(days=int(row['DLT']))
                else:
                    inv_min = adu * (1 + row['LTF'])
                    inv_max = inv_min + adu * row['DLT'] * (1 + row['VF'])
                    pedido = max(inv_max - posicion, 0)
                    fecha_entrega = fecha_eval + timedelta(days=int(row['DLT']))

                if pedido > 0:
                    pedidos_futuros.append((fecha_entrega, pedido))

                inventario_proy = max(0, stock_inicial + recepciones - consumo_semana)
                cobertura = inventario_proy / adu if adu else 0

                entry = {
                    'Referencia': ref,
                    'Mes': fecha.strftime('%b-%y'),
                    'Semana': semana,
                    'Fecha_Evaluacion': fecha_eval.strftime('%Y-%m-%d'),
                    'Recepciones_Planeadas': round(recepciones, 2),
                    'Transito': round(transito, 2),
                    'Stock_Inicial': round(stock_inicial, 2),
                    'Posicion': round(posicion, 2),
                    'Fecha_Pedido': fecha_eval.strftime('%Y-%m-%d'),
                    'Fecha_Entrega_Pedido': fecha_entrega.strftime('%Y-%m-%d'),
                    'Consumo_Proy': round(consumo_semana, 2),
                    'Cantidad_Pedir': round(pedido, 2),
                    'Inventario_Proy': round(inventario_proy, 2),
                    'Nivel_Cobertura_Dias': round(cobertura, 2),
                    'precio': round(precio, 2),
                    'vr_compra': round(precio * pedido, 2)
                }

                if gestion == 'Buffer':
                    entry.update({k: round(buf.get(k, 0), 2) for k in ['Red_Total', 'Yellow', 'Green', 'TO_Yellow', 'TO_Green']})
                else:
                    inv_obj = (inv_min + inv_max) / 2
                    entry.update({
                        'Inventario_Minimo': round(inv_min, 2),
                        'Inventario_Maximo': round(inv_max, 2),
                        'Inventario_Objetivo': round(inv_obj, 2)
                    })

                historico.append(entry)
                stock_inicial = inventario_proy

        resultados[ref] = (gestion, pd.DataFrame(historico))

    return df_params.set_index('Referencia'), resultados

def exportar_resumen(ruta_excel, ruta_maestro, salida):
    df_params, resultados = ejecutar_completo(ruta_excel, ruta_maestro)
    df_maestro = pd.read_excel(ruta_maestro, sheet_name='Codigo')
    df_maestro['Codigo'] = df_maestro['Codigo'].apply(normalizar_codigo)
    df_maestro = df_maestro.set_index('Codigo')[['Familia', 'Subfamilia', 'Descripcion_F']]

    wb = Workbook()
    wb.remove(wb.active)

    for categoria in ['Buffer', 'No Buffer']:
        datos = [df for ref, (gest, df) in resultados.items() if gest == categoria]
        if not datos:
            continue
        df_cat = pd.concat(datos, ignore_index=True)
        df_cat[['Familia', 'Subfamilia', 'Descripcion_F']] = df_cat['Referencia'].map(df_maestro.to_dict(orient='index')).apply(pd.Series)

        ws_detalle = wb.create_sheet(categoria)
        for r in dataframe_to_rows(df_cat, index=False, header=True):
            ws_detalle.append(r)

        df_resumen = df_cat[df_cat['vr_compra'] > 0].copy()
        df_resumen['Mes_Orden'] = pd.to_datetime(df_resumen['Mes'], format='%b-%y', errors='coerce')
        tabla = df_resumen.pivot_table(index=['Familia', 'Subfamilia', 'Referencia', 'Descripcion_F'],
                                       columns='Mes_Orden', values='vr_compra', aggfunc='sum', fill_value=0)
        tabla.columns = [d.strftime('%b-%y') for d in tabla.columns]
        tabla['Total'] = tabla.sum(axis=1)
        tabla = tabla.reset_index()

        total = pd.DataFrame([['TOTAL', '', '', ''] + list(tabla.iloc[:, 4:].sum())], columns=tabla.columns)
        tabla = pd.concat([tabla, total], ignore_index=True)

        ws_resumen = wb.create_sheet(f"Resumen_{categoria}")
        for r in dataframe_to_rows(tabla, index=False, header=True):
            ws_resumen.append(r)

    wb.save(salida)

if __name__ == '__main__':
    limpiar_precio_en_excel('Referencia V2.xlsx')
    exportar_resumen('Referencia V2.xlsx', 'data-consumo1.xlsx', 'Resumen_Buffer_NoBuffer_Semanal.xlsx')
    print("✅ Archivo generado correctamente.")
