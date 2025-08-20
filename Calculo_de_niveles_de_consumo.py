def ejecutar_parametros(archivo_consumo='data-consumo1.xlsx', archivo_referencias='Referencia V2.xlsx', fecha_final_str='26/06/2025'):
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    # === CARGA DE DATOS ===
    df_ltd = pd.read_excel(archivo_consumo, sheet_name="LTD")
    df_consumo = pd.read_excel(archivo_consumo, sheet_name="data-consumo")
    df_codigo = pd.read_excel(archivo_consumo, sheet_name="Codigo")
        # === REEMPLAZAR SOLO LA COLUMNA 'Codigo' EN LA HOJA 'ABC' CON CÓDIGOS DESDE HOJA 'Codigo' ===
    from openpyxl.utils import get_column_letter

        # === ACTUALIZAR SOLO LA COLUMNA 'Codigo' EN LA HOJA ABC ===
    wb_maestro = load_workbook(archivo_consumo)
    if 'ABC' in wb_maestro.sheetnames:
        ws_abc = wb_maestro['ABC']
        codigos_unicos = df_codigo['Codigo'].dropna().astype(str).str.strip().str.upper().str.replace('.0', '', regex=False).unique()

        header_row = [cell.value for cell in ws_abc[1]]
        try:
            col_codigo_idx = header_row.index('Codigo') + 1  # Excel usa índice desde 1
        except ValueError:
            raise ValueError("❌ No se encontró la columna 'Codigo' en la hoja ABC.")

        for i, cod in enumerate(sorted(codigos_unicos), start=2):  # desde fila 2 para evitar encabezado
            ws_abc.cell(row=i, column=col_codigo_idx, value=cod)

        # Limpiar códigos sobrantes (si antes había más filas)
        for j in range(i + 1, ws_abc.max_row + 1):
            ws_abc.cell(row=j, column=col_codigo_idx, value=None)

        wb_maestro.save(archivo_consumo)
    else:
        print("⚠️ La hoja 'ABC' no fue encontrada, no se actualizó la columna 'Codigo'.")

    df_abc = pd.read_excel(archivo_consumo, sheet_name="ABC", header=0)

    # === NORMALIZAR COLUMNAS ===
    df_abc.columns = [col.strip().upper() for col in df_abc.columns]
    if 'CODIGO' in df_abc.columns:
        df_abc = df_abc.rename(columns={'CODIGO': 'Codigo'})

    for df in [df_abc, df_codigo]:
        df['Codigo'] = df['Codigo'].astype(str).str.strip().str.upper().str.replace('.0', '', regex=False)

    # === FUNCIONES DE NORMALIZACIÓN ===
    def encontrar_columna(df, opciones):
        for col in df.columns:
            if col.lower().strip() in opciones:
                return col
        return None

    def normalizar_columnas_consumo(df):
        col_codigo = encontrar_columna(df, ['codigo', 'código', 'cod'])
        col_consumo = encontrar_columna(df, ['consumo'])
        col_fecha = encontrar_columna(df, ['fecha'])

        df = df.rename(columns={
            col_codigo: 'Codigo',
            col_consumo: 'Consumo',
            col_fecha: 'Fecha'
        })

        df['Codigo'] = df['Codigo'].astype(str).str.strip().str.upper().str.replace('.0', '', regex=False)
        df['Fecha'] = pd.to_datetime(df['Fecha'], errors='coerce', dayfirst=True)
        df['Consumo'] = pd.to_numeric(df['Consumo'], errors='coerce').fillna(0)
        return df

    def normalizar_columnas_ltd(df):
        col_codigo = encontrar_columna(df, ['codigo', 'código', 'cod'])
        col_mean_entrega = encontrar_columna(df, ['meanentrega', 'lt', 'leadtime'])
        col_moq = encontrar_columna(df, ['moq'])

        df = df.rename(columns={
            col_codigo: 'Codigo',
            col_mean_entrega: 'MeanEntrega',
            col_moq: 'MOQ'
        })

        df['Codigo'] = df['Codigo'].astype(str).str.strip().str.upper().str.replace('.0', '', regex=False)
        df['MeanEntrega'] = pd.to_numeric(df['MeanEntrega'], errors='coerce')
        df['MOQ'] = pd.to_numeric(df['MOQ'], errors='coerce')
        return df

    # === LIMPIEZA ===
    df_ltd = normalizar_columnas_ltd(df_ltd)
    df_consumo = normalizar_columnas_consumo(df_consumo)

    # === DICCIONARIOS ===
    ltd_dict = df_ltd.set_index('Codigo')['MeanEntrega'].to_dict()
    MOQ_dict = df_ltd.set_index('Codigo')['MOQ'].to_dict()
    subfamilia_dict = df_codigo.set_index('Codigo')['Subfamilia'].to_dict()
    perfil_dict = df_abc.set_index('Codigo')['PERFIL'].to_dict()

    # === PARÁMETROS DE FECHA ===
    fecha_final_real = pd.to_datetime(fecha_final_str, dayfirst=True)
    fecha_inicio_real = fecha_final_real - pd.Timedelta(days=179)

    # === PROCESAMIENTO ===
    resumen = []
    no_encontrados = []

    for codigo in df_consumo['Codigo'].unique():
        lt = ltd_dict.get(codigo)
        MOQ = MOQ_dict.get(codigo)

        if pd.isna(lt) or MOQ is None or pd.isna(MOQ):
            continue

        df_cod = df_consumo[df_consumo['Codigo'] == codigo].copy()
        df_cod = df_cod.groupby('Fecha', as_index=False).agg({'Consumo': 'sum'}).sort_values('Fecha')

        fechas_completas = pd.date_range(start=fecha_inicio_real, end=fecha_final_real, freq='D')
        df_cod = df_cod.set_index('Fecha').reindex(fechas_completas).fillna(0).reset_index()
        df_cod.columns = ['Fecha', 'Consumo']
        df_cod['Codigo'] = codigo

        fecha_inicio_minima = fecha_inicio_real + pd.Timedelta(days=int(lt) - 1)

        consumos_lt, fechas_validas = [], []
        for i in range(len(df_cod)):
            fecha_actual = df_cod.loc[i, 'Fecha']
            if fecha_actual < fecha_inicio_minima or fecha_actual > fecha_final_real:
                continue
            fecha_inicio_lt = fecha_actual - pd.Timedelta(days=int(lt) - 1)
            ventana = df_cod[(df_cod['Fecha'] >= fecha_inicio_lt) & (df_cod['Fecha'] <= fecha_actual)]
            suma_consumo = ventana['Consumo'].sum()
            consumos_lt.append(suma_consumo)
            fechas_validas.append(fecha_actual)

        if not consumos_lt:
            continue

        promedio = np.mean(consumos_lt)
        desviacion = np.std(consumos_lt, ddof=1) if len(consumos_lt) > 1 else 0

        ltf = 0.25 if lt <= 6 else 0.5 if lt <= 12 else 0.75
        adu = round(promedio / lt, 2) if lt else 0
        vf = round(desviacion / promedio, 2) if promedio else 0
        vf = min(vf, 1)

        try:
            oc_base = int(round(MOQ / adu)) if adu > 0 else 7
            order_cycle = 1 if oc_base == 0 else oc_base
        except:
            order_cycle = 7

        perfil = str(perfil_dict.get(codigo, '')).strip().upper()
        if perfil == '':
            no_encontrados.append(codigo)

        resumen.append({
            'Referencia': codigo,
            'ADU': adu,
            'LTF': ltf,
            'VF': vf,
            'MOQ': MOQ,
            'order_cycle': order_cycle,
            'DLT': int(lt),
            'Consumo_Promedio_LT': round(promedio, 2),
            'Desviacion_Estandar_LT': round(desviacion, 2),
            'Ventanas_Validas': len(consumos_lt),
            'Fecha_Inicio_VentanaCompleta': min(fechas_validas).strftime('%Y-%m-%d'),
            'Fecha_Fin_VentanaCompleta': max(fechas_validas).strftime('%Y-%m-%d'),
            'subfamilia': subfamilia_dict.get(codigo, 'SIN_SUBFAMILIA'),
            'gestion': 'Buffer' if perfil in ['A-X', 'A-Y', 'B-X'] else 'No Buffer'
        })

    # === EXPORTAR ===
    df_final = pd.DataFrame(resumen)

    # Guardar hoja "prueba" en data-consumo1.xlsx
    with pd.ExcelWriter(archivo_consumo, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        df_final.to_excel(writer, sheet_name='prueba', index=False)

    # Guardar hoja "Referencias" sin afectar las demás hojas
    wb = load_workbook(archivo_referencias)
    if 'Referencias' in wb.sheetnames:
        del wb['Referencias']
    ws = wb.create_sheet('Referencias')
    columnas_referencias = ['Referencia', 'ADU', 'LTF', 'VF', 'MOQ', 'order_cycle', 'DLT', 'gestion']
    for r in dataframe_to_rows(df_final[columnas_referencias], index=False, header=True):
        ws.append(r)
    wb.save(archivo_referencias)

    print("✅ Proceso completado.")
    if no_encontrados:
        print(f"❗ {len(no_encontrados)} códigos no tienen perfil en ABC. Ejemplos: {no_encontrados[:5]}")

if __name__ == "__main__":
    ejecutar_parametros()
